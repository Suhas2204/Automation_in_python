import openpyxl as xl
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference

filename = 'transaction_id.xlsx'
def process_name(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]
    sheet.delete_cols(4)
    bold_font = Font(bold=True)


    for row in range (2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        cell2 = cell.value
        corrected_price = (cell2 * 0.9)
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
        corrected_price_cell.font = bold_font


    values = Reference(sheet,
              min_row= 2,
              max_row= sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'a6')


    wb.save(filename)


